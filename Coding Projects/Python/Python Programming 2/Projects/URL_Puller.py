import openpyxl
import re

# Replace 'your_excel_file.xlsx' with the path to your Excel file
excel_file_path = 'APLUPresidents-OfficeMailingAddress.xlsx'

# Create a regular expression pattern to match URLs
url_pattern = r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'

# Open the Excel file
workbook = openpyxl.load_workbook(excel_file_path)

# Initialize a list to store extracted URLs
extracted_urls = []

# Loop through all sheets in the Excel file
for sheet in workbook.sheetnames:
    worksheet = workbook[sheet]

    # Iterate through all cells in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        for cell_value in row:
            if cell_value is not None and isinstance(cell_value, str):
                # Find URLs in the cell's text using the regular expression
                urls = re.findall(url_pattern, cell_value)
                extracted_urls.extend(urls)

# Print the extracted URLs
for url in extracted_urls:
    print(url)
